from selenium import webdriver
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait #ilgili driverı bekleten yapı
from selenium.webdriver.support import expected_conditions as ec #beklenen koşullar
from constants import globalConstants as c
import pytest
import openpyxl

     #Kullandığımız SauceDemo sitesinde kendi belirlediğiniz en az "3" test case daha yazınız.
     #En az 1 testiniz parametrize fonksiyonu ile en az 3 farklı veriyle test edilmelidir.
     
     #Case 1: Login sonrası açılan menü konrolü
     #Case 2: Ödeme bilgileri girişinin kontrolü
     #Case 3: Siparişin tamamlandığının kontrolü
     
class Test_DemoClass:   
    
    
    def setup_method(self): #her test başlangıcında çalışacak fonksiyon
        self.driver = webdriver.Chrome()
        self.driver.get(c.BASE_URL)
        self.driver.maximize_window() 
          

    def teardown_method(self): # her testinin bitiminde çalışacak fonksiyon
        self.driver.quit()
    

    def getData():
        excel=openpyxl.load_workbook(c.test_continue)
        sheet=excel["Sayfa1"] #hangi sayfada çalıştığımızı gösteriyor
        rows =sheet.max_row #kaçıncı satıra kadar veri var?
        data =[]
        
        for i in range(2,rows+1):
            firsname = sheet.cell(i,1).value
            lastname = sheet.cell(i,2).value
            postalcode = sheet.cell(i,3).value
            data.append((firsname,lastname,postalcode))
        
        return data
    
    
    #Case 1: (Açılan menü sonucu çıkan sayfada olduğumuzun kontrolü
    #Adim1:Username  password gir ve login ol
    #Adim2:Open menuye tikla
    #Adim3:Acilan sıralı listeden about a tikla
    #Beklenen sonuç:Sayfanin ortasinde yer alan  renksiz ve küçük  puntoyla yazilmis yaziyi kontrol et ve yazının bu olduğunu görmelisin:
    #The world relies on your code. Test on thousands of different device, browser, and OS configurations–anywhere, any time."
        
    def test_openmenu(self):    
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID , c.USERNAME_ID)))
        usernameInput.send_keys("standard_user")
        passwordInput =WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID , c.PASSWORD_ID)))
        passwordInput.send_keys("secret_sauce")
        loginButton = self.driver.find_element(By.ID ,c.LOGINBUTTON_ID)
        loginButton.click() 
        openMenu=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID ,"react-burger-menu-btn")))
        openMenu.click()
        about=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID ,"about_sidebar_link")))
        about.click()
        littleExplanation=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH ,"//*[@id='__next']/div[2]/div[1]/div/div[1]/div[1]/div/div[3]/p")))
        text=littleExplanation.text
        excepted_text="The world relies on your code. Test on thousands of different device, browser, and OS configurations–anywhere, any time."
        assert text== excepted_text
        
        
        
 
     #Case 2: Ödeme bilgileri girişinin kontrolü
     #Adim1:Username  password gir ve login ol 
     #Adım2:Sepete tıkla
     #Adım3:checkout tıkla     
     #Adım4: Firstname,lastname ve postal code girişlerini yap
     #Adım5: Continue butonuna tıkla
     #Beklenen Sonuç: (https://www.saucedemo.com/checkout-step-two.html ) bu URL de olduğunu görmelisin.      
    
    
    
    
    
    @pytest.mark.parametrize("firstname,lastname,postalcode",getData())
    def test_continue(self,firstname,lastname,postalcode): 
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID ,c.USERNAME_ID)))
        usernameInput.send_keys("standard_user")
        passwordInput =WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID , c.PASSWORD_ID)))
        passwordInput.send_keys("secret_sauce")
        loginButton = self.driver.find_element(By.ID ,c.LOGINBUTTON_ID)
        loginButton.click() 
        shoppingCart=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='shopping_cart_container']/a")))
        shoppingCart.click()
        checkOut=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.CHECKOUTBUTTON_ID)))
        checkOut.click()
        firstName=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.FIRSTNAME_ID)))
        firstName.send_keys(firstname)
        lastName=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.LASTNAME_ID)))
        lastName.send_keys(lastname)
        postalCode=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.POSTALCODE_ID)))
        postalCode.send_keys(postalcode)
        ctnButton =WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='continue']")))
        ctnButton.click()
        current_url = self.driver.current_url
        expected_url = "https://www.saucedemo.com/checkout-step-two.html"
        if (expected_url == current_url ):
            firstTest = True
        else:
            firstTest = False



     #Case 3: Siparişin tamamlandığının kontrolü
     #Adim1:Username  password gir ve login ol 
     #Adım2:Sepete tıkla
     #Adım3:checkout tıkla     
     #Adım4: Firstname,lastname ve postal code girişlerini yap
     #Adım5: Continue butonuna tıkla
     #Adım6: Finish butonuna tıkla       
     #Beklenen Sonuç: "Thank you for your order!"sayfanın ortasında görünmelidir.        
    def test_thankyou_page(self):
            
        usernameInput = WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID , c.USERNAME_ID)))
        usernameInput.send_keys("standard_user")
        passwordInput =WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID , c.PASSWORD_ID)))
        passwordInput.send_keys("secret_sauce")
        loginButton = self.driver.find_element(By.ID ,"login-button")
        loginButton.click() 
        shoppingCart=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='shopping_cart_container']/a")))
        shoppingCart.click()
        checkOut=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.CHECKOUTBUTTON_ID)))
        checkOut.click()
        firstName=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.FIRSTNAME_ID)))
        firstName.send_keys("Mehtap")
        lastName=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.FIRSTNAME_ID)))
        lastName.send_keys("Tunç")
        postalCode=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,c.POSTALCODE_ID)))
        postalCode.send_keys("77777")
        ctnButton =WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='continue']")))
        ctnButton.click()  
        finishButton=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.ID,"finish")))
        finishButton.click()
        thankyouText=WebDriverWait(self.driver,5).until(ec.visibility_of_element_located((By.XPATH,"//*[@id='checkout_complete_container']/h2")))
        text= thankyouText.text
        excepted_text="Thank you for your order!"
        assert text== excepted_text



        
        
# testClass=Test_DemoClass
# testClass.setup_method()
# testClass.teardown_method()
# testClass.test_continue()
# testClass.test_openmenu()
# testClass.test_thankyou_page()
                        

        
        
        

    

        

       


    
   








        
    
    
    